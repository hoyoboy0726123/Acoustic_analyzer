# -*- coding: utf-8 -*-
"""
聲學測試 AI 分析系統 - 報告生成模組 (AUD-008)

功能:
- 彙整各個分析模組的結果
- 生成 Excel 格式的測試報告 (含完整 Raw Data)
"""

import pandas as pd
import io
import os
from datetime import datetime
from typing import Optional, Dict, Any, Tuple
import numpy as np

# Import core analysis functions
from core.noise_level import calculate_noise_level
from core.band_analyzer import compute_octave_bands
from core.sop_analyzer import analyze_idle_mode, analyze_ue_mode, analyze_workload_mode
from core.fft import compute_average_spectrum, compute_peak_hold_spectrum, compute_psd
from core.discrete_tone import detect_discrete_tones

def generate_excel_report(
    audio_data: np.ndarray,
    sample_rate: int,
    filename: str = "audio.wav",
    sop_params: dict = None,
    calibration_offset: float = 0.0,
    analysis_settings: dict = None
) -> Tuple[Optional[bytes], Optional[str]]:
    """生成完整 Excel 測試報告
    
    Args:
        audio_data: 音訊數據
        sample_rate: 取樣率
        filename: 原始檔名
        sop_params: SOP 分析參數
        calibration_offset: 校準偏移 (dB)
        analysis_settings: 所有分析設定參數
        
    Returns:
        Tuple[bytes, str]: (Excel檔案二進位數據, 錯誤訊息/None)
    """
    try:
        # 預設分析設定
        if analysis_settings is None:
            analysis_settings = {}
        
        # 獲取設定參數
        use_a_weighting = analysis_settings.get('use_a_weighting', True)
        spectrum_mode = analysis_settings.get('spectrum_mode', 'average')
        window_function = analysis_settings.get('window_function', 'hann')
        n_fft = analysis_settings.get('n_fft', 8192)
        ecma_standard = analysis_settings.get('ecma_standard', 'ECMA-74')
        spectrogram_spl_offset = analysis_settings.get('spectrogram_spl_offset', 0.0)
        highpass_cutoff = analysis_settings.get('highpass_cutoff', 20)
        
        # 1. 計算全域噪音指標 (AUD-003) + Time Profile
        noise_metrics = calculate_noise_level(audio_data, sample_rate)
        
        # 2. 計算 1/3 倍頻程數據 (AUD-006)
        octave_data = compute_octave_bands(audio_data, sample_rate, use_a_weighting=use_a_weighting)
        
        # 3. ASUS SOP Analysis
        sop_result = None
        if sop_params:
            mode = sop_params.get('mode', 'IDLE')
            if mode == 'IDLE':
                sop_result = analyze_idle_mode(audio_data, sample_rate, sop_params.get('idle_spec', 20.0))
            elif mode == 'UE':
                sop_result = analyze_ue_mode(audio_data, sample_rate)
            elif mode == 'Workload':
                sop_result = analyze_workload_mode(audio_data, sample_rate, sop_params.get('work_spec_fail', 22.0), sop_params.get('work_spec_max', 28.0))
        
        # 4. 計算 FFT 頻譜 (所有模式)
        fft_avg_freqs, fft_avg_mags = compute_average_spectrum(audio_data, sample_rate, window=window_function, n_fft=n_fft)
        fft_peak_freqs, fft_peak_mags = compute_peak_hold_spectrum(audio_data, sample_rate, window=window_function, n_fft=n_fft)
        psd_freqs, psd_mags = compute_psd(audio_data, sample_rate, window=window_function, n_fft=n_fft)
        
        # 5. Discrete Tone 檢測
        discrete_tone_result = detect_discrete_tones(
            audio_data, sample_rate, 
            spectrum_mode=spectrum_mode, 
            window_function=window_function, 
            n_fft=n_fft,
            ecma_standard=ecma_standard
        )

        # --- 建構 DataFrame ---
        
        # ========== Sheet 1: Summary (完整設定) ==========
        duration = len(audio_data) / sample_rate
        
        status = "N/A"
        if sop_result:
            status = "PASS" if sop_result.get('is_pass', True) else "FAIL"
        
        # 計算頻率解析度
        freq_resolution = sample_rate / n_fft
            
        summary_rows = [
            # --- 檔案資訊 ---
            ("【檔案資訊】", ""),
            ("檔案名稱", filename),
            ("分析日期時間", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("音訊長度 (秒)", round(duration, 3)),
            ("取樣率 (Hz)", sample_rate),
            ("樣本數", len(audio_data)),
            ("位元深度", "16-bit (假設)"),
            
            # --- 校準設定 ---
            ("", ""),
            ("【校準設定】", ""),
            ("麥克風校準偏移 (dB)", calibration_offset),
            ("Spectrogram SPL 偏移 (dB)", spectrogram_spl_offset),
            
            # --- 分析設定 ---
            ("", ""),
            ("【分析設定】", ""),
            ("A-weighting 加權", "是" if use_a_weighting else "否"),
            ("頻譜模式", spectrum_mode),
            ("窗函數", window_function),
            ("FFT 點數 (N)", n_fft),
            ("頻率解析度 (Hz)", round(freq_resolution, 4)),
            ("高通濾波截止頻率 (Hz)", highpass_cutoff),
            ("ECMA 標準", ecma_standard),
            
            # --- 噪音指標 (校準後) ---
            ("", ""),
            ("【噪音指標 (已套用校準)】", ""),
            ("Leq (dBA)", round(noise_metrics['leq_dba'] + calibration_offset, 2)),
            ("Lmax (dBA)", round(noise_metrics['lmax_dba'] + calibration_offset, 2)),
            ("Lmin (dBA)", round(noise_metrics['lmin_dba'] + calibration_offset, 2)),
            ("L10 (dBA)", round(noise_metrics['l10'] + calibration_offset, 2)),
            ("L50 (dBA)", round(noise_metrics.get('l50', 0) + calibration_offset, 2)),
            ("L90 (dBA)", round(noise_metrics['l90'] + calibration_offset, 2)),
            
            # --- SOP 分析結果 ---
            ("", ""),
            ("【SOP 分析結果】", ""),
            ("SOP 模式", sop_params.get('mode', 'N/A') if sop_params else "N/A"),
            ("SOP 判定結果", status),
        ]
        
        if sop_result:
            if sop_params['mode'] == 'UE':
                summary_rows.append(("SOP 平均 Leq", round(sop_result['leq'] + calibration_offset, 2)))
            else:
                summary_rows.append(("SOP 最大 Leq", round(sop_result['max_leq'] + calibration_offset, 2)))
                if sop_params['mode'] == 'Workload':
                    summary_rows.append(("失敗率 (%)", sop_result['fail_rate']))
            
            summary_rows.append(("分析摘要", sop_result['summary']))
        
        # --- Discrete Tone 結果 ---
        summary_rows.append(("", ""))
        summary_rows.append(("【Discrete Tone 檢測結果】", ""))
        summary_rows.append(("檢測到 Discrete Tone", "是" if discrete_tone_result.get('tone_detected', False) else "否"))
        summary_rows.append(("Discrete Tone 數量", len(discrete_tone_result.get('tones', []))))
        
        if discrete_tone_result.get('tones'):
            for i, tone in enumerate(discrete_tone_result['tones']):
                summary_rows.append((f"Tone #{i+1} 頻率 (Hz)", round(tone.get('frequency', 0), 2)))
                summary_rows.append((f"Tone #{i+1} 幅度 (dB)", round(tone.get('magnitude_db', 0) + calibration_offset, 2)))
                summary_rows.append((f"Tone #{i+1} PR (dB)", round(tone.get('prominence', 0), 2)))
                summary_rows.append((f"Tone #{i+1} TNR (dB)", round(tone.get('tnr', 0), 2)))

        df_summary = pd.DataFrame(summary_rows, columns=["Parameter", "Value"])
        
        # ========== Sheet 2: 1/3 Octave Bands ==========
        df_octave = pd.DataFrame({
            "Center Frequency (Hz)": octave_data["nominal_freqs"],
            "Band Level dB(A)": [round(x + calibration_offset, 3) for x in octave_data["band_levels"]]
        })
        
        # ========== Sheet 3: Level vs Time (Full Profile) ==========
        profile = noise_metrics.get("profile", {})
        times_data = profile.get("times", [])
        levels_data = profile.get("levels", [])
        
        # 安全檢查：避免 numpy array 真值判斷錯誤
        has_profile_data = (
            profile is not None and 
            len(times_data) > 0 if hasattr(times_data, '__len__') else times_data is not None
        ) and (
            len(levels_data) > 0 if hasattr(levels_data, '__len__') else levels_data is not None
        )
        
        if has_profile_data:
            df_profile = pd.DataFrame({
                "Time (sec)": [round(float(t), 4) for t in times_data],
                "Level dB(A)": [round(float(l) + calibration_offset, 3) for l in levels_data]
            })
        else:
            df_profile = pd.DataFrame(columns=["Time (sec)", "Level dB(A)"])

        # ========== Sheet 4: SOP Moving Average ==========
        df_moving = None
        if sop_result and 'leqs' in sop_result:
            df_moving = pd.DataFrame({
                "Time (sec)": [round(t, 3) for t in sop_result['times']],
                "Moving Leq (dBA)": [round(l + calibration_offset, 3) for l in sop_result['leqs']]
            })

        # ========== Sheet 5: FFT Average Spectrum (完整輸出) ==========
        mask_avg = fft_avg_freqs <= 20000
        df_fft_avg = pd.DataFrame({
            "Frequency (Hz)": [round(f, 4) for f in fft_avg_freqs[mask_avg]],
            "Magnitude dB": [round(m + calibration_offset, 3) for m in fft_avg_mags[mask_avg]]
        })
        
        # ========== Sheet 6: FFT Peak Hold Spectrum (完整輸出) ==========
        mask_peak = fft_peak_freqs <= 20000
        df_fft_peak = pd.DataFrame({
            "Frequency (Hz)": [round(f, 4) for f in fft_peak_freqs[mask_peak]],
            "Magnitude dB": [round(m + calibration_offset, 3) for m in fft_peak_mags[mask_peak]]
        })
        
        # ========== Sheet 7: PSD Spectrum (完整輸出) ==========
        mask_psd = psd_freqs <= 20000
        df_psd = pd.DataFrame({
            "Frequency (Hz)": [round(f, 4) for f in psd_freqs[mask_psd]],
            "PSD dB/Hz": [round(m + calibration_offset, 3) for m in psd_mags[mask_psd]]
        })
        
        # ========== Sheet 8: Discrete Tone Raw Data ==========
        all_candidates = discrete_tone_result.get('all_candidates', [])
        if all_candidates:
            df_discrete_tone = pd.DataFrame({
                "Frequency (Hz)": [round(t.get('frequency', 0), 2) for t in all_candidates],
                "Magnitude dB": [round(t.get('magnitude_db', 0) + calibration_offset, 2) for t in all_candidates],
                "PR (dB)": [round(t.get('prominence', 0), 2) for t in all_candidates],
                "PR Threshold (dB)": [round(t.get('pr_threshold', 0), 2) for t in all_candidates],
                "TNR (dB)": [round(t.get('tnr', 0), 2) for t in all_candidates],
                "TNR Threshold (dB)": [round(t.get('tnr_threshold', 0), 2) for t in all_candidates],
                "Exceeds Threshold": [t.get('exceeds_threshold', False) for t in all_candidates],
                "Method": [t.get('method', 'N/A') for t in all_candidates],
                "Band": [t.get('band', 'N/A') for t in all_candidates]
            })
        else:
            df_discrete_tone = pd.DataFrame(columns=[
                "Frequency (Hz)", "Magnitude dB", "PR (dB)", "PR Threshold (dB)",
                "TNR (dB)", "TNR Threshold (dB)", "Exceeds Threshold", "Method", "Band"
            ])

        # --- 輸出到 Excel ---
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
            df_octave.to_excel(writer, sheet_name='1-3 Octave Bands', index=False)
            df_profile.to_excel(writer, sheet_name='Level vs Time', index=False)
            if df_moving is not None:
                df_moving.to_excel(writer, sheet_name='SOP Moving Avg', index=False)
            df_fft_avg.to_excel(writer, sheet_name='FFT Average', index=False)
            df_fft_peak.to_excel(writer, sheet_name='FFT Peak Hold', index=False)
            df_psd.to_excel(writer, sheet_name='PSD', index=False)
            df_discrete_tone.to_excel(writer, sheet_name='Discrete Tone', index=False)
            
            # 格式化 Summary sheet
            from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
            
            workbook = writer.book
            summary_sheet = workbook['Summary']
            
            # 定義樣式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            header_font = Font(bold=True, size=12)
            section_font = Font(bold=True, size=11, color='FFFFFF')
            section_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            value_alignment = Alignment(horizontal='left', vertical='center')
            
            # 設定欄寬
            summary_sheet.column_dimensions['A'].width = 35
            summary_sheet.column_dimensions['B'].width = 45
            
            # 套用樣式到所有儲存格
            for row in range(1, summary_sheet.max_row + 1):
                cell_a = summary_sheet.cell(row=row, column=1)
                cell_b = summary_sheet.cell(row=row, column=2)
                
                # 套用框線
                cell_a.border = thin_border
                cell_b.border = thin_border
                cell_a.alignment = value_alignment
                cell_b.alignment = value_alignment
                
                # 檢查是否為區段標題 (以【開頭)
                if cell_a.value and str(cell_a.value).startswith('【'):
                    cell_a.font = section_font
                    cell_a.fill = section_fill
                    cell_b.fill = section_fill
                # 標題列
                elif row == 1:
                    cell_a.font = header_font
                    cell_b.font = header_font
                    header_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
                    cell_a.fill = header_fill
                    cell_b.fill = header_fill
            
            # 格式化其他 sheet 的標題列
            for sheet_name in ['1-3 Octave Bands', 'Level vs Time', 'FFT Average', 'FFT Peak Hold', 'PSD', 'Discrete Tone']:
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    header_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
                    for col in range(1, sheet.max_column + 1):
                        cell = sheet.cell(row=1, column=col)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # 套用框線到資料區
                    for row in range(2, min(sheet.max_row + 1, 10002)):  # 限制以避免過慢
                        for col in range(1, sheet.max_column + 1):
                            sheet.cell(row=row, column=col).border = thin_border
            
        return output.getvalue(), None
        
    except Exception as e:
        import traceback
        return None, f"{str(e)}\n{traceback.format_exc()}"

